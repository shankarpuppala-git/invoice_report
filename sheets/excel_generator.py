"""
sheets/excel_generator.py
──────────────────────────
Builds the Invoice Report Excel workbook in memory for any tenant.
Returns BytesIO — nothing written to disk.

Email column is populated from auth_users.user_id (resolved via bulk lookup
in report_service), NOT from pso.notif_email.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from service.report_service import ReportResult

from common.logger import get_logger

logger = get_logger(__name__)

# ─── Palette ─────────────────────────────────────────────────────────────────
HEADER_BG  = "1A1A2E"
HEADER_FG  = "FFFFFF"
TITLE_FG   = "1A1A2E"
SETTLED_FG = "1E6B2E"
VOIDED_FG  = "8B0000"
ALT_ROW_BG = "F2F2F2"

COLUMNS = [
    ("Order No.",          18),
    ("Ordered Date",       26),
    ("Order Total($)",     18),
    ("Invoice Total($)",   18),
    ("Transaction ID",     22),
    ("Transaction Status", 26),
    ("Invoice No.",        18),
    ("Invoiced Date",      26),
    ("Customer No.",       18),
    ("Email",              36),   # ← populated from auth_users.user_id
]

SUMMARY_ROWS = [
    "Total number of orders placed using credit card payments",
    "Total number of orders successfully invoiced",
    "Total number of invoiced orders for which payment was successfully settled",
    "Total number of orders where payment was voided or cancelled prior to settlement",
    "Total number of orders with payments not yet settled as of the report date",
]


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M:%S UTC")
    return str(value)


def _fmt_currency(value) -> str:
    if value is None:
        return ""
    try:
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return ""


# ─── Main builder ─────────────────────────────────────────────────────────────

def generate_excel(result: "ReportResult") -> BytesIO:
    """
    Builds the workbook in memory and returns a seeked BytesIO.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    total_cols      = len(COLUMNS)
    last_col_letter = get_column_letter(total_cols)

    # ── Row 1: Title ──────────────────────────────────────────────
    title = (
        f"{result.tenant_name} - Invoice Report  Summary "
        f"{result.start_date.strftime('%m/%d/%Y')} - "
        f"{result.end_date.strftime('%m/%d/%Y')}"
    )
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(bold=True, size=13, color=TITLE_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: blank ──────────────────────────────────────────────
    ws.append([])

    # ── Rows 3-7: Summary metrics ─────────────────────────────────
    summary_values = [
        result.total_credit_card_orders,
        result.total_invoiced,
        result.total_settled,
        result.total_voided,
        result.total_pending,
    ]
    for label, value in zip(SUMMARY_ROWS, summary_values):
        row_num = ws.max_row + 1
        ws.append([None])
        ws.row_dimensions[row_num].height = 16

        ws.merge_cells(f"A{row_num}:E{row_num}")
        lc = ws[f"A{row_num}"]
        lc.value     = label
        lc.font      = Font(size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws[f"F{row_num}"]
        vc.value     = value
        vc.font      = Font(bold=True, size=10)
        vc.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 8: blank ──────────────────────────────────────────────
    ws.append([])

    # ── Row 9: Column headers ─────────────────────────────────────
    header_row   = ws.max_row + 1
    header_fill  = PatternFill("solid", fgColor=HEADER_BG)
    header_font  = Font(bold=True, color=HEADER_FG, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell            = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = header_align
        cell.border     = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[header_row].height = 22

    # ── Data rows ─────────────────────────────────────────────────
    alt_fill     = PatternFill("solid", fgColor=ALT_ROW_BG)
    settled_font = Font(color=SETTLED_FG, size=10)
    voided_font  = Font(color=VOIDED_FG,  size=10)
    default_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    for row_idx, order in enumerate(result.orders):
        data_row = header_row + row_idx + 1

        row_values = [
            order.ext_order_number,
            _fmt_dt(order.ordered_date),
            _fmt_currency(order.order_total),
            _fmt_currency(order.invoiced_amount),
            order.payment_reference_no or "",
            order.transaction_status,
            order.invoice_number or "",
            _fmt_dt(order.invoiced_date),
            order.customer_number or "",
            order.email or "",          # ← from auth_users.user_id bulk lookup
        ]

        row_fill     = alt_fill if row_idx % 2 == 1 else None
        status_lower = (order.transaction_status or "").lower()

        if status_lower == "settledsuccessfully":
            status_font = settled_font
        elif status_lower in ("voided", "refundsettledsuccessfully"):
            status_font = voided_font
        else:
            status_font = default_font

        for col_idx, value in enumerate(row_values, start=1):
            cell           = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border    = _thin_border()
            cell.font      = status_font if col_idx == 6 else default_font
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[data_row].height = 16

    ws.freeze_panes  = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    logger.info(
        "Excel built | rows=%d | size=%d KB",
        len(result.orders),
        len(excel_io.getvalue()) // 1024,
    )
    return excel_io